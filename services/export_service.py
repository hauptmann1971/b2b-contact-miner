from sqlalchemy.orm import Session
from models.database import DomainContact, Contact, ContactType, Keyword, SearchResult
from typing import List, Optional
from datetime import datetime
import csv
import io
from loguru import logger


class ExportService:
    def __init__(self, db: Session):
        self.db = db
    
    def export_to_flat_csv(self, filters: dict = None) -> str:
        """Export contacts to flat CSV format with keyword info"""
        contacts = self._query_contacts_with_keyword(filters)
        
        output = io.StringIO()
        writer = csv.writer(output, delimiter=';')
        
        # Headers in Russian
        writer.writerow([
            "№",
            "Ключевое слово",
            "Страна",
            "Язык",
            "Домен",
            "Email",
            "Telegram",
            "LinkedIn",
            "Телефон",
            "Предметная область"
        ])
        
        row_number = 1
        for contact_data in contacts:
            writer.writerow([
                row_number,
                contact_data['keyword'],
                contact_data['country'],
                contact_data['language'],
                contact_data['domain'],
                contact_data['email'],
                contact_data['telegram'],
                contact_data['linkedin'],
                contact_data['phone'],
                contact_data.get('subject_area', '')
            ])
            row_number += 1
        
        logger.info(f"Exported {row_number - 1} contacts to flat CSV")
        
        # Add BOM for Excel compatibility
        return '\ufeff' + output.getvalue()
    
    def _query_contacts_with_keyword(self, filters: dict = None) -> List[dict]:
        """Query contacts joined with keyword and search result info"""
        query = (
            self.db.query(
                Keyword.keyword,
                Keyword.country,
                Keyword.language,
                DomainContact.domain,
                DomainContact.tags,
                Contact.contact_type,
                Contact.value
            )
            .join(SearchResult, DomainContact.search_result_id == SearchResult.id)
            .join(Keyword, SearchResult.keyword_id == Keyword.id)
            .join(Contact, Contact.domain_contact_id == DomainContact.id)
        )
        
        if filters:
            if "keyword_id" in filters:
                query = query.filter(Keyword.id == filters["keyword_id"])
            
            if "country" in filters:
                query = query.filter(Keyword.country == filters["country"])
            
            if "contact_type" in filters:
                query = query.filter(Contact.contact_type == filters["contact_type"])
            
            if "domain" in filters:
                query = query.filter(DomainContact.domain.ilike(f"%{filters['domain']}%"))
        
        results = query.order_by(Keyword.id, DomainContact.domain).all()
        
        # Group by domain to combine all contact types
        domain_groups = {}
        for row in results:
            key = (row.keyword, row.country, row.language, row.domain)
            if key not in domain_groups:
                domain_groups[key] = {
                    'keyword': row.keyword,
                    'country': row.country,
                    'language': row.language,
                    'domain': row.domain,
                    'emails': [],
                    'telegrams': [],
                    'linkedins': [],
                    'phones': [],
                    'tags': row.tags or []
                }
            
            if row.contact_type == ContactType.EMAIL:
                domain_groups[key]['emails'].append(row.value)
            elif row.contact_type == ContactType.TELEGRAM:
                domain_groups[key]['telegrams'].append(row.value)
            elif row.contact_type == ContactType.LINKEDIN:
                domain_groups[key]['linkedins'].append(row.value)
            elif row.contact_type == ContactType.PHONE:
                domain_groups[key]['phones'].append(row.value)
        
        # Convert to list of dicts
        flat_contacts = []
        for key, data in domain_groups.items():
            # Create one row per domain with all contacts
            flat_contacts.append({
                'keyword': data['keyword'],
                'country': data['country'],
                'language': data['language'],
                'domain': data['domain'],
                'email': '; '.join(data['emails']),
                'telegram': '; '.join(data['telegrams']),
                'linkedin': '; '.join(data['linkedins']),
                'phone': '; '.join(data['phones']),
                'subject_area': self._extract_subject_area(data['tags'])
            })
        
        return flat_contacts
    
    def _extract_subject_area(self, tags: list) -> str:
        """Extract subject area from tags"""
        if not tags or not isinstance(tags, list):
            return ''
        
        # Filter out generic tags and keep only meaningful categories
        generic_tags = {'b2b', 'company', 'business', 'website'}
        meaningful_tags = [tag for tag in tags if tag.lower() not in generic_tags]
        
        return ', '.join(meaningful_tags[:3])  # Top 3 tags
    
    def export_to_csv(self, filters: dict = None) -> str:
        """Export contacts to CSV format"""
        contacts = self._query_contacts(filters)
        
        output = io.StringIO()
        writer = csv.writer(output)
        
        writer.writerow([
            "Domain",
            "Email",
            "Telegram",
            "LinkedIn",
            "Phone",
            "Confidence Score",
            "Extraction Method",
            "Tags",
            "Created At"
        ])
        
        domain_contacts = {}
        for contact in contacts:
            domain = contact.domain_contact.domain
            if domain not in domain_contacts:
                domain_contacts[domain] = {
                    "emails": [],
                    "telegrams": [],
                    "linkedins": [],
                    "phones": [],
                    "confidence": contact.domain_contact.confidence_score,
                    "method": contact.domain_contact.extraction_method,
                    "tags": contact.domain_contact.tags,
                    "created_at": contact.domain_contact.created_at
                }
            
            if contact.contact_type == ContactType.EMAIL:
                domain_contacts[domain]["emails"].append(contact.value)
            elif contact.contact_type == ContactType.TELEGRAM:
                domain_contacts[domain]["telegrams"].append(contact.value)
            elif contact.contact_type == ContactType.LINKEDIN:
                domain_contacts[domain]["linkedins"].append(contact.value)
            elif contact.contact_type == ContactType.PHONE:
                domain_contacts[domain]["phones"].append(contact.value)
        
        for domain, data in domain_contacts.items():
            writer.writerow([
                domain,
                "; ".join(data["emails"]),
                "; ".join(data["telegrams"]),
                "; ".join(data["linkedins"]),
                "; ".join(data["phones"]),
                data["confidence"],
                data["method"],
                "; ".join(data["tags"]) if isinstance(data["tags"], list) else data["tags"],
                data["created_at"].strftime("%Y-%m-%d %H:%M:%S")
            ])
        
        logger.info(f"Exported {len(domain_contacts)} domains to CSV")
        return output.getvalue()
    
    def export_to_excel(self, filters: dict = None) -> bytes:
        """Export contacts to Excel format"""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            contacts = self._query_contacts(filters)
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Contacts"
            
            headers = [
                "Domain", "Email", "Telegram", "LinkedIn", "Phone",
                "Confidence", "Method", "Tags", "Created At"
            ]
            ws.append(headers)
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            domain_contacts = {}
            for contact in contacts:
                domain = contact.domain_contact.domain
                if domain not in domain_contacts:
                    domain_contacts[domain] = {
                        "emails": [],
                        "telegrams": [],
                        "linkedins": [],
                        "phones": [],
                        "confidence": contact.domain_contact.confidence_score,
                        "method": contact.domain_contact.extraction_method,
                        "tags": contact.domain_contact.tags,
                        "created_at": contact.domain_contact.created_at
                    }
                
                if contact.contact_type == ContactType.EMAIL:
                    domain_contacts[domain]["emails"].append(contact.value)
                elif contact.contact_type == ContactType.TELEGRAM:
                    domain_contacts[domain]["telegrams"].append(contact.value)
                elif contact.contact_type == ContactType.LINKEDIN:
                    domain_contacts[domain]["linkedins"].append(contact.value)
                elif contact.contact_type == ContactType.PHONE:
                    domain_contacts[domain]["phones"].append(contact.value)
            
            for domain, data in domain_contacts.items():
                ws.append([
                    domain,
                    "\n".join(data["emails"]),
                    "\n".join(data["telegrams"]),
                    "\n".join(data["linkedins"]),
                    "\n".join(data["phones"]),
                    data["confidence"],
                    data["method"],
                    ", ".join(data["tags"]) if isinstance(data["tags"], list) else data["tags"],
                    data["created_at"].strftime("%Y-%m-%d %H:%M:%S")
                ])
            
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            logger.info(f"Exported {len(domain_contacts)} domains to Excel")
            return output.getvalue()
            
        except ImportError:
            logger.error("openpyxl not installed. Install with: pip install openpyxl")
            raise
    
    def _query_contacts(self, filters: dict = None) -> List[Contact]:
        """Query contacts with optional filters"""
        query = self.db.query(Contact).join(DomainContact)
        
        if filters:
            if "min_confidence" in filters:
                query = query.filter(DomainContact.confidence_score >= filters["min_confidence"])
            
            if "contact_type" in filters:
                query = query.filter(Contact.contact_type == filters["contact_type"])
            
            if "domain" in filters:
                query = query.filter(DomainContact.domain.ilike(f"%{filters['domain']}%"))
            
            if "start_date" in filters:
                query = query.filter(DomainContact.created_at >= filters["start_date"])
            if "end_date" in filters:
                query = query.filter(DomainContact.created_at <= filters["end_date"])
            
            if "keyword_id" in filters:
                query = query.join(Contact.domain_contact).join(DomainContact.search_result).filter(
                    Contact.domain_contact.has(
                        DomainContact.search_result.has(keyword_id=filters["keyword_id"])
                    )
                )
        
        return query.order_by(DomainContact.confidence_score.desc()).all()
    
    def get_export_summary(self) -> dict:
        """Get summary statistics for export"""
        total_domains = self.db.query(DomainContact.domain).distinct().count()
        total_emails = self.db.query(Contact).filter(Contact.contact_type == ContactType.EMAIL).count()
        total_telegram = self.db.query(Contact).filter(Contact.contact_type == ContactType.TELEGRAM).count()
        total_linkedin = self.db.query(Contact).filter(Contact.contact_type == ContactType.LINKEDIN).count()
        
        return {
            "total_domains": total_domains,
            "total_emails": total_emails,
            "total_telegram": total_telegram,
            "total_linkedin": total_linkedin,
            "total_contacts": total_emails + total_telegram + total_linkedin,
            "export_generated_at": datetime.utcnow().isoformat()
        }
